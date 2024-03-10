import asyncio
from datetime import date
from random import shuffle
from string import ascii_lowercase

from aiogram.utils.markdown import hbold
from bot.buttons.text import mission_message
from db.model import User, Unit, Vocabulary, Test, TestSection, EverestVocab
import aiohttp
import openpyxl
from aiogram.utils.markdown import hlink , hcode
from aiogram.enums.parse_mode import ParseMode

admins = [1148477816,6308227858]


def is_admin(msg):
    return msg.from_user.id in admins

async def check_url(url):
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as response:
                return response.status == 200
    except:
        return False




async def read_excel(filename, unit_id):
    wb_obj = openpyxl.load_workbook(filename)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    data = dict()
    for i in range(1,max_row+1):
        eng = sheet_obj.cell(row=i, column=1).value
        uzb = sheet_obj.cell(row=i, column=2).value
        await Vocabulary.create(eng = eng , uzb = uzb, unit_id= unit_id)

async def everest_vocab_excel(filename):
    wb_obj = openpyxl.load_workbook(filename)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    for i in range(1,max_row+1):
        eng = sheet_obj.cell(row=i, column=1).value
        uzb = sheet_obj.cell(row=i, column=2).value
        await EverestVocab.create(eng = eng , uzb = uzb)


async def create_text_vocab(vocabularies):
    text = ""
    for vocab in vocabularies:
        text += f"{vocab.eng}  ".ljust(10) + f"{vocab.uzb}\n"
    return text


async def today_yesterday(users):
    today, yesterday = 0, 0
    for user in users:
        day = user.created_at.date()
        if day == date.today():
            today += 1
        elif (date.today() - day).days == 1:
            yesterday += 1
    return today, yesterday


async def send_message(msg, mission, users):
    un_block, block = 0, 0
    for user in users:
        try:
            if un_block % 5 == 1:
                await mission.edit_text(text=hbold(mission_message.format(un_block, block)))
            await msg.copy_to(user.id)
            un_block += 1
        except:
            block += 1
    await mission.edit_text(text=hbold(mission_message.format(un_block, block)))


async def read_document_test(file, test_section_id):
    wb_obj = openpyxl.load_workbook(file)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    # try:
    for i in range(1, max_row + 1):
        question = sheet_obj.cell(row=i, column=1).value
        a = str(sheet_obj.cell(row=i, column=2).value)
        b = str(sheet_obj.cell(row=i, column=3).value)
        c = str(sheet_obj.cell(row=i, column=4).value)
        d = str(sheet_obj.cell(row=i, column=5).value)
        right = sheet_obj.cell(row=i, column=6).value
        await Test.create(question =question , a =a , b =b , c =c , d=d ,right=right, test_section_id= test_section_id)
    # except:
    #     return False
    return True

async def shuffle_vocab(click_units):
    vocab_list = []
    for unit in click_units:
        vocabs = await Vocabulary.get_unit_id(unit)
        vocab_list.extend(vocabs)
    shuffle(vocab_list)
    return vocab_list[:20]

async def shuffle_test(test_section):
    test_list = []
    for test in test_section:
        vocabs = await Test.get_test_section_id(test)
        test_list.extend(vocabs)
    shuffle(test_list)
    return test_list[:10]





def write_text(vocabularies , book_icon):
    text = ""
    for vocab in vocabularies:
        text += f"{book_icon}{hbold(vocab.eng)}-{vocab.uzb}\n"
    return text


def text_shablon_test(test):
    text = f"{hbold(test.question)}\n"
    text += hbold(f"a) {test.a}\n")
    text += hbold(f"b) {test.b}\n")
    text += hbold(f"c) {test.c}\n")
    text += hbold(f"d) {test.d}\n")
    return text




async def write_text_for_go_vocab(click_unit):
    text = ""
    for unit in click_unit:
        unit = await Unit.get(unit)
        text += f"Unit {unit.unit}\n"
    text += "👆 Quydagi unitlardan random 20 ta savol beriladi ♻️"
    return text

async def write_text_for_go_test(test_section):
    text = ""
    for test in test_section:
        test = await TestSection.get(test)
        text += hbold(f"{test.title}\n")
    text += "👆 Quydagi testlardan random 10 ta savol beriladi ♻️"
    return text

async def send_to_admin(msg , text):
    for admin in admins:
        await msg.bot.send_message(admin, text)

async def vocab_test_result(msg, right , wrong, type = 'Everest Vocab'):

    ability = right * 100 //  (wrong+right)
    text = f"""{hbold(msg.from_user.full_name)})
📌 Natija  
🟢 to'g'ri: {right}
🔴 xatolar: {wrong}
📊 o'zlashtirish: {ability}%"""
    admin_text = f"""t.me/{msg.from_user.username}
📌 Natija : {type}
🟢 to'g'ri: {right}
🔴 xatolar: {wrong}
📊 o'zlashtirish: {ability}%"""
    await send_to_admin(msg , admin_text)

    return text

async def test_result_text(call, right , wrong):

    ability = right * 100 //  (right+wrong)
    text = f"""{hbold(call.from_user.full_name)})
📌 {hcode("TEST")} Natijasi  
🟢 to'g'ri: {right}
🔴 xatolar: {wrong}
📊 o'zlashtirish: {ability}%"""
    admin_text = f"""t.me/{call.from_user.username} | {hcode("TEST")}
📌 Natija
🟢 to'g'ri: {right}
🔴 xatolar: {wrong}
📊 o'zlashtirish: {ability}%"""
    await send_to_admin(call.message , admin_text)
    return text
